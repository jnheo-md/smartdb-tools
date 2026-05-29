#!/usr/bin/env python3
"""
Build a cross-hospital SmartDB field reference dataset.

The output is intended to be committed to the repository under
reference/hospital-field-reference/ so MCP clients can fetch the latest
GitHub-hosted JSON cache. It stores schema/layout/value-encoding metadata and
aggregate data-presence indicators only; it does not export patient-level data.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import httpx

SESSION_DIR = Path.home() / ".smartdb"
SESSION_FILE = SESSION_DIR / "session.json"
CONFIG_FILE = SESSION_DIR / "config.json"
DEFAULT_API_URL = "https://api.ai.smartstroke.net"
DEFAULT_OUTPUT_DIR = Path("reference/hospital-field-reference")
DEFAULT_REFERENCE_URL = (
    "https://raw.githubusercontent.com/jnheo-md/smartdb-tools/master/"
    "reference/hospital-field-reference/smartdb_field_reference.json"
)


class APIError(RuntimeError):
    def __init__(self, message: str, status_code: int | None = None):
        super().__init__(message)
        self.status_code = status_code


class SmartDBClient:
    def __init__(self, api_url: str | None = None):
        self.api_url = (api_url or get_api_url()).rstrip("/")
        self.client = httpx.Client(timeout=90.0, headers=self._headers())

    def close(self) -> None:
        self.client.close()

    def get(self, path: str, params: dict[str, Any] | None = None) -> Any:
        response = self.client.get(self.api_url + path, params=params)
        return self._json_or_error(response)

    def post(self, path: str, json_body: dict[str, Any] | None = None) -> Any:
        response = self.client.post(self.api_url + path, json=json_body)
        return self._json_or_error(response)

    def _headers(self) -> dict[str, str]:
        headers = {"Accept": "application/json"}
        token = get_token()
        if token:
            headers["Authorization"] = f"Bearer {token}"
        return headers

    def _json_or_error(self, response: httpx.Response) -> Any:
        if response.status_code < 400:
            return response.json()

        detail = f"API error (HTTP {response.status_code})"
        try:
            body = response.json()
            detail = body.get("detail", detail)
        except Exception:
            pass
        raise APIError(detail, status_code=response.status_code)


def get_api_url() -> str:
    url = os.environ.get("SMARTDB_API_URL")
    if url:
        return url.rstrip("/")
    if CONFIG_FILE.exists():
        try:
            data = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
            if data.get("api_url"):
                return str(data["api_url"]).rstrip("/")
        except (json.JSONDecodeError, OSError):
            pass
    if SESSION_FILE.exists():
        try:
            data = json.loads(SESSION_FILE.read_text(encoding="utf-8"))
            if data.get("api_url"):
                return str(data["api_url"]).rstrip("/")
        except (json.JSONDecodeError, OSError):
            pass
    return DEFAULT_API_URL


def get_token() -> str | None:
    if not SESSION_FILE.exists():
        return None
    try:
        data = json.loads(SESSION_FILE.read_text(encoding="utf-8"))
        return data.get("access_token")
    except (json.JSONDecodeError, OSError):
        return None


def stable_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def short_hash(value: Any) -> str:
    return hashlib.sha256(stable_json(value).encode("utf-8")).hexdigest()[:16]


def normalize_value_map(value_map: Any) -> dict[str, str]:
    if not isinstance(value_map, dict):
        return {}
    return {
        str(key): str(value)
        for key, value in sorted(value_map.items(), key=lambda item: str(item[0]))
    }


def sanitize_count(
    count: int | None,
    privacy: str,
    min_cell_size: int,
) -> tuple[int | None, str, bool | None]:
    if count is None:
        return None, "unknown", None
    if privacy == "internal":
        return count, str(count), count > 0
    if count == 0:
        return None, "0", False
    if count < min_cell_size:
        return None, f"1-{min_cell_size - 1}", True
    return None, f"{min_cell_size}+", True


def list_hospitals(client: SmartDBClient, selected_codes: set[str] | None) -> list[dict[str, Any]]:
    hospitals = client.get("/schema/hospitals")
    normalized = []
    for hospital in hospitals:
        code = str(hospital.get("code", ""))
        if selected_codes and code not in selected_codes:
            continue
        normalized.append({
            "code": code,
            "hidx": hospital.get("hidx"),
            "name": hospital.get("name", ""),
            "root_tables": hospital.get("root_tables", []),
        })
    return normalized


def get_table_names(client: SmartDBClient, hospital_code: str) -> list[str]:
    tables = client.get(f"/schema/tables/{hospital_code}")
    return [str(table.get("table", "")) for table in tables if table.get("table")]


def get_layout_variables(
    client: SmartDBClient,
    hospital_code: str,
    table_name: str,
) -> tuple[list[dict[str, Any]], str]:
    variables: dict[str, dict[str, Any]] = {}
    try:
        sections = client.get(f"/schema/sections/{hospital_code}/{table_name}")
    except APIError:
        sections = []

    if sections:
        for order in sorted({section["section_order"] for section in sections}):
            try:
                section_data = client.get(
                    f"/schema/section-vars/{hospital_code}/{table_name}/{order}"
                )
            except APIError:
                continue
            for subsection in section_data.get("subsections", []):
                for variable in subsection.get("variables", []):
                    key = variable.get("key")
                    if key:
                        variables[str(key)] = variable
        return list(variables.values()), "form_layout"

    try:
        table_vars = client.get(f"/schema/table-vars/{hospital_code}/{table_name}")
    except APIError:
        table_vars = []
    for variable in table_vars:
        key = variable.get("key")
        if key:
            variables[str(key)] = variable
    return list(variables.values()), "table_vars"


def get_count(
    client: SmartDBClient,
    hospital_code: str,
    filters: list[dict[str, Any]],
) -> tuple[int | None, str]:
    try:
        result = client.post(
            "/query/count",
            json_body={"hospital": hospital_code, "filters": filters},
        )
        return int(result.get("count", 0)), ""
    except APIError as exc:
        status = f"HTTP {exc.status_code}: " if exc.status_code else ""
        return None, f"{status}{exc}"


def build_field_record(
    hospital: dict[str, Any],
    table_name: str,
    layout_source: str,
    layout_variable: dict[str, Any],
    variable_info: dict[str, Any],
    total_count: int | None,
    total_count_bucket: str,
    non_null_count: int | None,
    non_null_count_bucket: str,
    has_data: bool | None,
    count_error: str,
) -> dict[str, Any]:
    value_map = normalize_value_map(variable_info.get("value_map"))
    options = str(variable_info.get("options", "") or "")
    type_payload = {
        "type": str(variable_info.get("type", "")),
        "type_label": str(variable_info.get("type_label", "")),
    }
    value_payload = {
        "type": str(variable_info.get("type", "")),
        "type_label": str(variable_info.get("type_label", "")),
        "options": options,
        "value_map": value_map,
    }
    key = str(variable_info.get("key") or layout_variable.get("key") or "")
    label = str(variable_info.get("label") or layout_variable.get("label") or "")

    return {
        "hospital": hospital["code"],
        "hidx": hospital.get("hidx"),
        "hospital_name": hospital.get("name", ""),
        "table": str(variable_info.get("table") or table_name),
        "field_key": key,
        "column": str(variable_info.get("col", "")),
        "label": label,
        "type": str(variable_info.get("type", "")),
        "type_label": str(variable_info.get("type_label", "")),
        "options": options,
        "value_map": value_map,
        "layout_source": layout_source,
        "type_signature": short_hash(type_payload),
        "value_signature": short_hash(value_payload),
        "label_signature": short_hash(label),
        "total_records": total_count,
        "total_records_bucket": total_count_bucket,
        "non_null_records": non_null_count,
        "non_null_records_bucket": non_null_count_bucket,
        "has_data": has_data,
        "count_error": count_error,
    }


def collect_fields(
    client: SmartDBClient,
    hospitals: list[dict[str, Any]],
    check_data: bool,
    privacy: str,
    min_cell_size: int,
    max_fields_per_hospital: int | None,
    workers: int,
) -> list[dict[str, Any]]:
    fields: list[dict[str, Any]] = []
    for hospital in hospitals:
        hospital_code = hospital["code"]
        print(f"Collecting {hospital_code}...", file=sys.stderr)
        hospital_fields: list[dict[str, Any]] = []
        total_raw = None
        total_bucket = "unknown"
        if check_data:
            total_raw_count, _ = get_count(client, hospital_code, [])
            total_raw, total_bucket, _ = sanitize_count(
                total_raw_count,
                privacy,
                min_cell_size,
            )

        hospital_field_count = 0
        for table_name in get_table_names(client, hospital_code):
            layout_variables, layout_source = get_layout_variables(
                client,
                hospital_code,
                table_name,
            )
            for layout_variable in layout_variables:
                field_key = str(layout_variable.get("key", ""))
                if not field_key:
                    continue
                if (
                    layout_variable.get("table")
                    and (
                        "value_map" in layout_variable
                        or "values" in layout_variable
                        or "options" in layout_variable
                    )
                ):
                    variable_info = dict(layout_variable)
                else:
                    try:
                        variable_info = client.get(
                            f"/schema/variable/{hospital_code}/{field_key}"
                        )
                    except APIError:
                        variable_info = {
                            "key": field_key,
                            "table": table_name,
                            "label": layout_variable.get("label", ""),
                            "type": layout_variable.get("type", ""),
                            "type_label": layout_variable.get("type_label", ""),
                            "options": layout_variable.get("options", ""),
                            "value_map": layout_variable.get("value_map", {}),
                        }

                hospital_fields.append(
                    build_field_record(
                        hospital=hospital,
                        table_name=table_name,
                        layout_source=layout_source,
                        layout_variable=layout_variable,
                        variable_info=variable_info,
                        total_count=total_raw,
                        total_count_bucket=total_bucket,
                        non_null_count=None,
                        non_null_count_bucket="not_checked",
                        has_data=None,
                        count_error="",
                    )
                )
                hospital_field_count += 1
                if max_fields_per_hospital and hospital_field_count >= max_fields_per_hospital:
                    break
            if max_fields_per_hospital and hospital_field_count >= max_fields_per_hospital:
                break

        if check_data and hospital_fields:
            print(
                f"  Checking aggregate data presence for {len(hospital_fields)} fields...",
                file=sys.stderr,
            )

            def count_one(index: int, field_key: str) -> tuple[int, int | None, str, bool | None, str]:
                raw_count, error = get_count(
                    client,
                    hospital_code,
                    [{"variable": field_key, "operator": "IS NOT NULL"}],
                )
                raw, bucket, has_data = sanitize_count(raw_count, privacy, min_cell_size)
                return index, raw, bucket, has_data, error

            max_workers = max(1, min(workers, 32))
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = [
                    executor.submit(count_one, index, field["field_key"])
                    for index, field in enumerate(hospital_fields)
                ]
                for completed, future in enumerate(as_completed(futures), 1):
                    index, raw, bucket, has_data, error = future.result()
                    hospital_fields[index]["non_null_records"] = raw
                    hospital_fields[index]["non_null_records_bucket"] = bucket
                    hospital_fields[index]["has_data"] = has_data
                    hospital_fields[index]["count_error"] = error
                    if completed % 500 == 0:
                        print(
                            f"    {completed}/{len(hospital_fields)} fields checked",
                            file=sys.stderr,
                        )

        fields.extend(hospital_fields)
    return fields


def add_difference(
    differences: list[dict[str, Any]],
    field_key: str,
    issue_type: str,
    severity: str,
    summary: str,
    present_hospitals: list[str],
    missing_hospitals: list[str],
    details: Any,
) -> None:
    differences.append({
        "field_key": field_key,
        "issue_type": issue_type,
        "severity": severity,
        "summary": summary,
        "present_hospitals": present_hospitals,
        "missing_hospitals": missing_hospitals,
        "details": details,
    })


def group_by(items: list[dict[str, Any]], key: str) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in items:
        grouped.setdefault(str(item.get(key, "")), []).append(item)
    return grouped


def compute_differences(
    fields: list[dict[str, Any]],
    hospitals: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    all_hospital_codes = [hospital["code"] for hospital in hospitals]
    differences: list[dict[str, Any]] = []
    fields_by_key = group_by(fields, "field_key")

    for field_key, field_records in sorted(fields_by_key.items()):
        present_hospitals = sorted({record["hospital"] for record in field_records})
        missing_hospitals = [
            hospital_code
            for hospital_code in all_hospital_codes
            if hospital_code not in present_hospitals
        ]
        if missing_hospitals:
            add_difference(
                differences,
                field_key,
                "missing_column",
                "decision_required",
                "Field is present in some hospitals but missing in others.",
                present_hospitals,
                missing_hospitals,
                {
                    record["hospital"]: {
                        "table": record["table"],
                        "label": record["label"],
                        "type_label": record["type_label"],
                    }
                    for record in field_records
                },
            )

        table_groups = group_by(field_records, "table")
        if len(table_groups) > 1:
            add_difference(
                differences,
                field_key,
                "table_location_differs",
                "review",
                "Field is stored in different table names across hospitals.",
                present_hospitals,
                missing_hospitals,
                {table: sorted(record["hospital"] for record in records) for table, records in table_groups.items()},
            )

        type_groups = group_by(field_records, "type_signature")
        if len(type_groups) > 1:
            add_difference(
                differences,
                field_key,
                "type_differs",
                "decision_required",
                "Field type differs across hospitals.",
                present_hospitals,
                missing_hospitals,
                {
                    signature: {
                        "hospitals": sorted(record["hospital"] for record in records),
                        "type_label": records[0]["type_label"],
                        "type": records[0]["type"],
                    }
                    for signature, records in type_groups.items()
                },
            )

        value_groups = group_by(field_records, "value_signature")
        if len(value_groups) > 1:
            add_difference(
                differences,
                field_key,
                "value_encoding_differs",
                "decision_required",
                "Field options/value map differs across hospitals.",
                present_hospitals,
                missing_hospitals,
                {
                    signature: {
                        "hospitals": sorted(record["hospital"] for record in records),
                        "type_label": records[0]["type_label"],
                        "options": records[0]["options"],
                        "value_map": records[0]["value_map"],
                    }
                    for signature, records in value_groups.items()
                },
            )

        label_groups = group_by(field_records, "label_signature")
        if len(label_groups) > 1:
            add_difference(
                differences,
                field_key,
                "label_differs",
                "review",
                "Field label differs across hospitals.",
                present_hospitals,
                missing_hospitals,
                {
                    signature: {
                        "hospitals": sorted(record["hospital"] for record in records),
                        "label": records[0]["label"],
                    }
                    for signature, records in label_groups.items()
                },
            )

        if any(record["has_data"] is not None for record in field_records):
            empty_hospitals = sorted(
                record["hospital"] for record in field_records if record["has_data"] is False
            )
            populated_hospitals = sorted(
                record["hospital"] for record in field_records if record["has_data"] is True
            )
            if empty_hospitals and populated_hospitals:
                add_difference(
                    differences,
                    field_key,
                    "data_presence_differs",
                    "decision_required",
                    "Field has data in some hospitals but no non-null records in others.",
                    present_hospitals,
                    missing_hospitals,
                    {
                        "populated_hospitals": populated_hospitals,
                        "empty_hospitals": empty_hospitals,
                    },
                )

    return differences


def build_field_index(
    fields: list[dict[str, Any]],
    differences: list[dict[str, Any]],
) -> dict[str, Any]:
    index: dict[str, Any] = {}
    for field in fields:
        field_key = field["field_key"]
        entry = index.setdefault(field_key, {"hospitals": {}, "differences": []})
        entry["hospitals"][field["hospital"]] = {
            "table": field["table"],
            "label": field["label"],
            "type": field["type"],
            "type_label": field["type_label"],
            "options": field["options"],
            "value_map": field["value_map"],
            "non_null_records": field["non_null_records"],
            "non_null_records_bucket": field["non_null_records_bucket"],
            "has_data": field["has_data"],
            "count_error": field["count_error"],
        }
    for difference in differences:
        index.setdefault(difference["field_key"], {"hospitals": {}, "differences": []})
        index[difference["field_key"]]["differences"].append(difference)
    return index


def write_json(dataset: dict[str, Any], output_dir: Path) -> Path:
    output_path = output_dir / "smartdb_field_reference.json"
    output_path.write_text(
        json.dumps(dataset, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return output_path


def write_csv_fallback(dataset: dict[str, Any], output_dir: Path) -> list[Path]:
    paths: list[Path] = []
    for name, rows in (
        ("fields", dataset["fields"]),
        ("differences", dataset["differences"]),
        ("hospitals", dataset["hospitals"]),
    ):
        path = output_dir / f"smartdb_field_reference_{name}.csv"
        if not rows:
            continue
        headers = sorted({key for row in rows for key in row.keys()})
        with path.open("w", newline="", encoding="utf-8-sig") as csv_file:
            writer = csv.DictWriter(csv_file, fieldnames=headers)
            writer.writeheader()
            for row in rows:
                writer.writerow({
                    key: stable_json(value) if isinstance(value, (dict, list)) else value
                    for key, value in row.items()
                })
        paths.append(path)
    return paths


def write_xlsx(dataset: dict[str, Any], output_dir: Path) -> Path | None:
    try:
        from openpyxl import Workbook
    except ImportError:
        return None

    output_path = output_dir / "smartdb_field_reference.xlsx"
    workbook = Workbook(write_only=True)

    metadata_sheet = workbook.create_sheet("Metadata")
    metadata = dataset["metadata"]
    metadata_sheet.append(["key", "value"])
    for key, value in metadata.items():
        metadata_sheet.append([key, stable_json(value) if isinstance(value, (dict, list)) else value])

    hospitals_sheet = workbook.create_sheet("Hospitals")
    hospitals_sheet.append(["code", "hidx", "name", "root_tables"])
    for hospital in dataset["hospitals"]:
        hospitals_sheet.append([
            hospital.get("code"),
            hospital.get("hidx"),
            hospital.get("name"),
            ", ".join(hospital.get("root_tables", [])),
        ])

    field_headers = [
        "hospital",
        "hidx",
        "hospital_name",
        "table",
        "field_key",
        "column",
        "label",
        "type",
        "type_label",
        "options",
        "value_map",
        "layout_source",
        "total_records",
        "total_records_bucket",
        "non_null_records",
        "non_null_records_bucket",
        "has_data",
        "count_error",
    ]
    fields_sheet = workbook.create_sheet("Fields")
    fields_sheet.append(field_headers)
    for field in dataset["fields"]:
        fields_sheet.append([
            stable_json(field.get(header)) if header == "value_map" else field.get(header)
            for header in field_headers
        ])

    diff_headers = [
        "field_key",
        "issue_type",
        "severity",
        "summary",
        "present_hospitals",
        "missing_hospitals",
        "details",
    ]
    differences_sheet = workbook.create_sheet("Differences")
    differences_sheet.append(diff_headers)
    for difference in dataset["differences"]:
        differences_sheet.append([
            stable_json(difference.get(header))
            if isinstance(difference.get(header), (dict, list))
            else difference.get(header)
            for header in diff_headers
        ])

    hospital_codes = [hospital["code"] for hospital in dataset["hospitals"]]
    presence_sheet = workbook.create_sheet("PresenceMatrix")
    presence_sheet.append(["field_key", "issue_types", *hospital_codes])
    differences_by_field: dict[str, list[str]] = {}
    for difference in dataset["differences"]:
        differences_by_field.setdefault(difference["field_key"], []).append(difference["issue_type"])
    for field_key, entry in sorted(dataset["field_index"].items()):
        row = [field_key, ", ".join(sorted(set(differences_by_field.get(field_key, []))))]
        for hospital_code in hospital_codes:
            hospital_field = entry["hospitals"].get(hospital_code)
            if not hospital_field:
                row.append("")
                continue
            row.append(
                f"{hospital_field['table']} | {hospital_field['type_label']} | "
                f"data={hospital_field['non_null_records_bucket']}"
            )
        presence_sheet.append(row)

    value_map_sheet = workbook.create_sheet("ValueMaps")
    value_map_sheet.append(["field_key", "hospital", "stored_value", "label"])
    for field in dataset["fields"]:
        for stored_value, label in field.get("value_map", {}).items():
            value_map_sheet.append([field["field_key"], field["hospital"], stored_value, label])

    workbook.save(output_path)
    return output_path


def write_readme(output_dir: Path, dataset: dict[str, Any]) -> Path:
    metadata = dataset["metadata"]
    readme_path = output_dir / "README.md"
    readme_path.write_text(
        "\n".join([
            "# SmartDB Hospital Field Reference",
            "",
            "Generated cross-hospital field reference artifacts for MCP and AI workflows.",
            "",
            "Files:",
            "- `smartdb_field_reference.json` — canonical machine-readable cache used by MCP.",
            "- `smartdb_field_reference.xlsx` — human review workbook with field presence and differences.",
            "",
            "Safety:",
            "- No patient-level rows are exported.",
            "- Default `public` privacy stores only data-presence buckets, not exact non-null counts.",
            "- Use `--privacy internal` only for private repositories or local-only caches.",
            "",
            "Regenerate:",
            "```bash",
            "python scripts/build_hospital_field_reference.py --privacy public",
            "```",
            "",
            f"Generated at: `{metadata['generated_at']}`",
            f"Privacy mode: `{metadata['privacy']}`",
            f"Hospitals: `{metadata['hospital_count']}`",
            f"Fields: `{metadata['field_count']}`",
            f"Differences: `{metadata['difference_count']}`",
            "",
        ]),
        encoding="utf-8",
    )
    return readme_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a GitHub-ready cross-hospital SmartDB field reference dataset.",
    )
    parser.add_argument(
        "--hospitals",
        default="",
        help="Comma-separated hospital codes. Default: all accessible hospitals.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help=f"Output directory. Default: {DEFAULT_OUTPUT_DIR}",
    )
    parser.add_argument(
        "--privacy",
        choices=["public", "internal"],
        default="public",
        help="public suppresses exact aggregate counts; internal stores exact aggregate counts.",
    )
    parser.add_argument(
        "--min-cell-size",
        type=int,
        default=5,
        help="Minimum public bucket size for non-zero counts. Default: 5.",
    )
    parser.add_argument(
        "--no-data-checks",
        action="store_true",
        help="Skip non-null count checks. Faster, but cannot flag empty fields.",
    )
    parser.add_argument(
        "--no-xlsx",
        action="store_true",
        help="Skip XLSX generation.",
    )
    parser.add_argument(
        "--max-fields-per-hospital",
        type=int,
        default=0,
        help="Testing/debug only: limit fields collected per hospital.",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=8,
        help="Parallel workers for aggregate data-presence checks. Default: 8.",
    )
    parser.add_argument(
        "--reference-url",
        default=DEFAULT_REFERENCE_URL,
        help="Raw GitHub URL where the JSON will be hosted.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    selected_codes = {
        code.strip()
        for code in args.hospitals.split(",")
        if code.strip()
    } or None
    output_dir = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    if args.privacy == "public":
        print(
            "Using public privacy mode: exact aggregate counts will be suppressed.",
            file=sys.stderr,
        )

    client = SmartDBClient()
    try:
        hospitals = list_hospitals(client, selected_codes)
        if not hospitals:
            raise APIError("No hospitals found or selected.")

        fields = collect_fields(
            client=client,
            hospitals=hospitals,
            check_data=not args.no_data_checks,
            privacy=args.privacy,
            min_cell_size=max(2, args.min_cell_size),
            max_fields_per_hospital=args.max_fields_per_hospital or None,
            workers=args.workers,
        )
    finally:
        client.close()

    differences = compute_differences(fields, hospitals)
    dataset = {
        "metadata": {
            "schema_version": 1,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "api_url": client.api_url,
            "reference_url": args.reference_url,
            "privacy": args.privacy,
            "data_checks": not args.no_data_checks,
            "min_cell_size": max(2, args.min_cell_size),
            "hospital_count": len(hospitals),
            "field_count": len(fields),
            "difference_count": len(differences),
            "notes": [
                "No patient-level data is included.",
                "value_map/options describe configured saved-value encodings, not observed patient values.",
            ],
        },
        "hospitals": hospitals,
        "fields": fields,
        "differences": differences,
        "field_index": build_field_index(fields, differences),
    }

    written_paths = [write_json(dataset, output_dir), write_readme(output_dir, dataset)]
    if not args.no_xlsx:
        xlsx_path = write_xlsx(dataset, output_dir)
        if xlsx_path:
            written_paths.append(xlsx_path)
        else:
            written_paths.extend(write_csv_fallback(dataset, output_dir))
            print(
                "openpyxl is not installed; wrote CSV fallback files instead of XLSX.",
                file=sys.stderr,
            )

    print("Wrote SmartDB field reference artifacts:")
    for path in written_paths:
        print(f"  {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
